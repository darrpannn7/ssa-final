import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ── colour palette ──────────────────────────────────────────────────
CLR = {
    "hdr_dark":   "0D1117",
    "hdr_blue":   "1565C0",
    "hdr_green":  "1B5E20",
    "hdr_purple": "4A148C",
    "hdr_teal":   "004D40",
    "hdr_orange": "E65100",
    "hdr_red":    "B71C1C",
    "hdr_indigo": "1A237E",
    "hdr_gray":   "37474F",
    "pass_bg":    "E8F5E9",
    "fail_bg":    "FFEBEE",
    "pend_bg":    "FFF8E1",
    "col_th":     "E3F2FD",
    "white":      "FFFFFF",
}

def mk_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mk_font(bold=False, color="000000", sz=10):
    return Font(bold=bold, color=color, size=sz, name="Calibri")

def mk_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def mk_align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_header_row(ws, row, values, bg, fg="FFFFFF", sz=11):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = mk_fill(bg)
        c.font = mk_font(bold=True, color=fg, sz=sz)
        c.alignment = mk_align("center")
        c.border = mk_border()

def write_section_title(ws, row, ncols, title, bg):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.fill = mk_fill(bg)
    c.font = mk_font(bold=True, color="FFFFFF", sz=12)
    c.alignment = mk_align("center")
    c.border = mk_border()

def write_row(ws, row, values, bg=None, bold=False):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        if bg:
            c.fill = mk_fill(bg)
        c.font = mk_font(bold=bold)
        c.alignment = mk_align()
        c.border = mk_border()

# ── COLUMNS shared across most sheets ───────────────────────────────
COLS = ["TC ID", "Section", "Test Case Title", "Preconditions",
        "Test Steps", "Expected Result", "Actual Result",
        "Status", "Priority", "Notes"]
COL_W = [10, 18, 30, 28, 42, 35, 22, 10, 10, 22]


# ═══════════════════════════════════════════════════════════════════
#  SHEET 1 – INDEX / COVERAGE MAP
# ═══════════════════════════════════════════════════════════════════
ws_idx = wb.create_sheet("📋 Index")
ws_idx.sheet_properties.tabColor = "0D1117"
set_col_widths(ws_idx, [6, 26, 40, 16, 16])

ws_idx.merge_cells("A1:E1")
c = ws_idx["A1"]
c.value = "SSA Analytics Dashboard – Test Case Coverage Index"
c.fill = mk_fill("0D1117"); c.font = mk_font(True, "FFFFFF", 14)
c.alignment = mk_align("center")

write_header_row(ws_idx, 2, ["#", "Sheet / Module", "Scope", "Total TCs", "Status"],
                 "1565C0")

idx_data = [
    (1,  "📋 Index",                "This sheet – coverage overview",                    "–",   "–"),
    (2,  "🖥 Backend – API Routes",  "All REST API endpoints",                             "28",  "Pending"),
    (3,  "⚙ Services – SunPy",      "Magnetogram download, caching, analysis",            "12",  "Pending"),
    (4,  "☁ Services – NOAA",       "GOES X-ray flux fetcher",                            "8",   "Pending"),
    (5,  "🛸 Services – CME",        "CME metadata, impact probability, LASCO image",      "10",  "Pending"),
    (6,  "💨 Services – Solar Wind", "Solar wind + IMF data fetching & processing",        "10",  "Pending"),
    (7,  "☢ Services – SEP",         "SEP flux, alerts, radiation risk calculation",       "12",  "Pending"),
    (8,  "🤖 AI – Inference",        "Chat endpoint, flare-predict endpoint",              "10",  "Pending"),
    (9,  "🌐 Frontend – Pages",      "Solar Flare, CME, SEP, Solar Wind pages",           "20",  "Pending"),
    (10, "🧩 Frontend – Components", "GOESFluxChart, FlareLog, CMECards, ServiceCards",   "14",  "Pending"),
    (11, "🔗 Frontend – API lib",    "lib/api.ts fetch functions",                        "10",  "Pending"),
    (12, "🔒 Non-Functional",        "Performance, security, accessibility, error handling","14", "Pending"),
    (13, "🔄 Integration E2E",       "End-to-end flows across FE ↔ BE ↔ external APIs",  "10",  "Pending"),
]

for r, row in enumerate(idx_data, 3):
    bg = "F5F5F5" if r % 2 == 0 else "FFFFFF"
    write_row(ws_idx, r, list(row), bg)

# ═══════════════════════════════════════════════════════════════════
#  helper: build a standard test-case sheet
# ═══════════════════════════════════════════════════════════════════
def build_tc_sheet(name, tab_color, header_color, sections):
    """
    sections = list of (section_title, [tc_rows])
    tc_row   = (tc_id, section, title, precond, steps, expected, actual, status, priority, notes)
    """
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_color
    set_col_widths(ws, COL_W)
    ws.row_dimensions[1].height = 30

    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    c = ws["A1"]
    c.value = name.split(" ", 1)[-1] + " – Test Cases"
    c.fill = mk_fill("0D1117"); c.font = mk_font(True, "FFFFFF", 13)
    c.alignment = mk_align("center")

    cur_row = 2
    for sec_title, rows in sections:
        write_section_title(ws, cur_row, len(COLS), sec_title, header_color)
        cur_row += 1
        write_header_row(ws, cur_row, COLS, header_color)
        cur_row += 1
        for i, r in enumerate(rows):
            bg = "F9F9F9" if i % 2 == 0 else "FFFFFF"
            write_row(ws, cur_row, list(r), bg)
            ws.row_dimensions[cur_row].height = 55
            cur_row += 1
        cur_row += 1  # blank separator
    return ws


# ═══════════════════════════════════════════════════════════════════
#  SHEET 2 – BACKEND API ROUTES
# ═══════════════════════════════════════════════════════════════════
PEND = "Pending"
api_sections = [
    ("Root & System Endpoints", [
        ("BE-001","Root","GET / returns running message","Backend running","GET http://localhost:8000/","Response 200 with {message: 'SSA Backend running successfully'}","","Pending","High",""),
        ("BE-002","System","GET /system/status returns health info","Backend running","GET /system/status","Response 200; contains service, time, magnetogram_cached, cache_location fields","","Pending","High",""),
        ("BE-003","System","magnetogram_cached is true after first magnetogram hit","Backend running; /magnetogram/latest called first","(1) Call /space-weather/magnetogram/latest\n(2) Call /system/status","magnetogram_cached = true","","Pending","Medium",""),
    ]),
    ("Solar Flare – Magnetogram Endpoints", [
        ("BE-004","Magnetogram","GET /magnetogram/latest returns success + data","Backend running","GET /space-weather/magnetogram/latest","status='success'; result contains data (list), meta.instrument='SDO/HMI', meta.unit='Gauss'","","Pending","High","First call downloads FITS – may be slow"),
        ("BE-005","Magnetogram","GET /magnetogram/image returns PNG","magnetogram/latest called","GET /space-weather/magnetogram/image","HTTP 200; Content-Type image/png; non-empty body","","Pending","High",""),
        ("BE-006","Magnetogram","GET /magnetogram/regions returns active regions","magnetogram/latest called","GET /space-weather/magnetogram/regions","status='success'; regions is an array; each region has id, bbox, strength, area, flare keys","","Pending","Medium",""),
        ("BE-007","Magnetogram","Cached FITS re-used within 30 min","First request made < 30 min ago","Make second request to /magnetogram/latest","Response time significantly faster; no new download","","Pending","Medium",""),
    ]),
    ("Solar Flare – Flares & AIA Endpoints", [
        ("BE-008","Solar Flares","GET /space-weather/flares returns list","Valid NASA API key configured","GET /space-weather/flares","200 OK; array of objects with classType, startTime, peakTime, endTime, activeRegion","","Pending","High",""),
        ("BE-009","Solar Flares","Flare data covers last 30 days","Backend running","GET /space-weather/flares","All items have startTime within last 30 days","","Pending","Medium",""),
        ("BE-010","AIA","GET /aia-image?wavelength=0171 returns JPEG","SDO server reachable","GET /space-weather/aia-image?wavelength=0171","200 OK; Content-Type image/jpeg","","Pending","High",""),
        ("BE-011","AIA","Invalid wavelength returns 400","Backend running","GET /space-weather/aia-image?wavelength=9999","400 Bad Request; detail lists valid wavelengths","","Pending","High",""),
        ("BE-012","AIA","Each valid wavelength returns image (0094,0131,0171,0193,0211,0304,0335,1600,1700)","Backend running","Loop GET for each wavelength","All return 200 OK with image/jpeg","","Pending","Medium",""),
    ]),
    ("GOES X-ray Endpoint", [
        ("BE-013","GOES","GET /noaa/goes-xray returns primary & secondary","NOAA SWPC reachable","GET /noaa/goes-xray","JSON with primary (list) and secondary (list); each item has time_tag and flux","","Pending","High",""),
        ("BE-014","GOES","primary list has ≤200 items","Backend running","GET /noaa/goes-xray","len(primary) <= 200","","Pending","Low",""),
        ("BE-015","GOES","flux values are positive numbers","Backend running","GET /noaa/goes-xray; inspect flux field","All flux values > 0","","Pending","Medium",""),
    ]),
    ("CME Endpoints", [
        ("BE-016","CME","GET /cme/full returns up to 10 events","NASA API key valid","GET /space-weather/cme/full","status='success'; cme_events array with activityID, startTime, speed, impactProbability","","Pending","High",""),
        ("BE-017","CME","GET /cme/image returns GIF","SOHO server reachable","GET /space-weather/cme/image","200 OK; Content-Type image/gif or application/octet-stream","","Pending","High",""),
        ("BE-018","CME","impactProbability is one of Low/Moderate/High","Backend running","GET /cme/full; inspect each cme_event.impactProbability","Only 'Low','Moderate','High' values present","","Pending","High",""),
    ]),
    ("Solar Wind Endpoints", [
        ("BE-019","Solar Wind","GET /wind/speed returns speed, density, temperature","NOAA SWPC reachable","GET /space-weather/wind/speed","status='success'; data list with time_tag, speed, density, temperature","","Pending","High",""),
        ("BE-020","Solar Wind","GET /wind/imf returns bx,by,bz,bt","NOAA SWPC reachable","GET /space-weather/wind/imf","status='success'; data list with bx, by, bz, bt","","Pending","High",""),
        ("BE-021","Solar Wind","GET /wind/all returns both solar_wind and imf","NOAA SWPC reachable","GET /space-weather/wind/all","status='success'; solar_wind list non-empty, imf list non-empty","","Pending","High",""),
        ("BE-022","Solar Wind","503 returned when NOAA unreachable","Simulate NOAA outage (mock)","GET /wind/speed with NOAA down","503 Service Unavailable with descriptive detail","","Pending","Medium",""),
    ]),
    ("SEP Endpoints", [
        ("BE-023","SEP","GET /sep/particle-flux returns proton and electron lists","NOAA SWPC reachable","GET /space-weather/sep/particle-flux","status='success'; proton and electron arrays with time_tag, flux, energy, satellite","","Pending","High",""),
        ("BE-024","SEP","GET /sep/alerts returns risk_level and alerts","NOAA alerts endpoint reachable","GET /space-weather/sep/alerts","status='success'; risk_level is one of quiet/low/moderate/high/severe; alerts is a list","","Pending","High",""),
        ("BE-025","SEP","GET /sep/all returns particle_flux, alerts, radiation_risk","NOAA reachable","GET /space-weather/sep/all","All three keys present; radiation_risk has crew, satellite, deep_space keys","","Pending","High",""),
        ("BE-026","SEP","radiation_risk values are valid levels","Backend running","GET /sep/all; check radiation_risk","Each of crew/satellite/deep_space is one of low/moderate/high/severe/extreme","","Pending","Medium",""),
    ]),
    ("AI Inference Endpoints", [
        ("BE-027","AI","POST /ai/chat with text-only message returns response","Backend running","POST /ai/chat with message='What is a solar flare?'","200 OK; JSON with response (str), surya_data, source, annotated_image=null, regions=[]","","Pending","High",""),
        ("BE-028","AI","POST /ai/flare-predict with use_live_data=true returns prediction","Backend running; live data reachable","POST /ai/flare-predict {use_live_data:true}","200 OK; predicted_class, confidence, onset_window_minutes, reasoning, model_source, timestamp","","Pending","High",""),
    ]),
]

build_tc_sheet("🖥 Backend – API Routes", "1565C0", "1565C0", api_sections)


# ═══════════════════════════════════════════════════════════════════
#  SHEET 3 – SUNPY PROCESSOR SERVICE
# ═══════════════════════════════════════════════════════════════════
sunpy_sections = [
    ("SunPyProcessor – _get_latest_hmi_url", [
        ("SP-001","SunPy","Scrapes today's JSOC directory and returns a .fits URL","Internet access to jsoc.stanford.edu","Call _get_latest_hmi_url()","Returns string ending in .fits","","Pending","High",""),
        ("SP-002","SunPy","Returns None when JSOC directory is unreachable","Simulate network timeout","Mock requests.get to raise Timeout; call _get_latest_hmi_url()","Returns None gracefully","","Pending","Medium",""),
    ]),
    ("SunPyProcessor – get_latest_magnetogram", [
        ("SP-003","SunPy","Downloads FITS on first call","No local cache exists","Call get_latest_magnetogram()","Returns dict with 'data' (list of lists) and 'meta' (date, instrument, unit)","","Pending","High",""),
        ("SP-004","SunPy","Uses cached FITS within 30 minutes","Cache file < 30 min old","Call get_latest_magnetogram() twice within 30 min","Second call prints 'Loading cached' not 'Fetching new'","","Pending","Medium",""),
        ("SP-005","SunPy","Data is clipped to ±150 Gauss","Any FITS file","Inspect max & min of returned data list","max(data) ≤ 150, min(data) ≥ -150","","Pending","High",""),
        ("SP-006","SunPy","Returned data shape matches 512×512 after resample","Any FITS file","len(data) == 512, len(data[0]) == 512","Both dimensions are 512","","Pending","Medium",""),
    ]),
    ("SunPyProcessor – analyze_magnetogram", [
        ("SP-007","SunPy","Returns mean_field, gradient_strength, polarity_mix","Data array available","Call analyze_magnetogram(data)","Dict with 3 float keys; mean_field ≥ 0","","Pending","Medium",""),
        ("SP-008","SunPy","polarity_mix is in range [0, 1]","Typical magnetogram data","Call analyze_magnetogram(data); inspect polarity_mix","0 ≤ polarity_mix ≤ 1","","Pending","Medium",""),
    ]),
    ("SunPyProcessor – calculate_flare_probability", [
        ("SP-009","SunPy","High strength+area returns high X-class probability","No precond","calculate_flare_probability(150, 3000)","X key ≥ 15","","Pending","High",""),
        ("SP-010","SunPy","Low strength+area returns high A/B probability","No precond","calculate_flare_probability(10, 50)","A + B ≥ 70","","Pending","Medium",""),
        ("SP-011","SunPy","Probabilities sum to 100","Any valid inputs","Sum A+B+C+M+X for any output","Sum == 100","","Pending","High",""),
    ]),
    ("SunPyProcessor – detect_active_regions", [
        ("SP-012","SunPy","Returns at most 8 regions","Any magnetogram data","Call detect_active_regions(data)","len(regions) ≤ 8","","Pending","Medium",""),
    ]),
]
build_tc_sheet("⚙ Services – SunPy", "00897B", "00897B", sunpy_sections)


# ═══════════════════════════════════════════════════════════════════
#  SHEET 4 – NOAA SERVICE
# ═══════════════════════════════════════════════════════════════════
noaa_sections = [
    ("NOAAFetcher – get_goes_xray_flux", [
        ("NO-001","NOAA","Fetches both GOES-16 and GOES-17 concurrently","NOAA SWPC reachable","await NOAAFetcher.get_goes_xray_flux()","Returns dict with primary and secondary; both non-empty lists","","Pending","High",""),
        ("NO-002","NOAA","Filters to 0.1-0.8nm long channel only","Backend running","Inspect energy field of returned items","All items have energy == '0.1-0.8nm'","","Pending","High",""),
        ("NO-003","NOAA","Flux values are positive","Backend running","Inspect flux field of all items","All flux > 0","","Pending","High",""),
        ("NO-004","NOAA","Returns ≤ 200 data points","Backend running","len(primary) and len(secondary)","Both ≤ 200","","Pending","Medium",""),
        ("NO-005","NOAA","Primary outage returns empty primary list, secondary still works","Simulate primary URL failure","Mock primary GET to raise; secondary normal","primary == [], secondary is non-empty list","","Pending","Medium",""),
        ("NO-006","NOAA","Both outages return empty lists","Simulate both URLs failing","Mock both GETs to raise","primary == [], secondary == []","","Pending","Medium",""),
        ("NO-007","NOAA","items have time_tag and flux keys only","Backend running","Inspect keys of each returned item","Keys are exactly {time_tag, flux}","","Pending","Low",""),
        ("NO-008","NOAA","time_tag is parseable ISO datetime string","Backend running","Parse each time_tag with datetime.fromisoformat","No parse errors","","Pending","Low",""),
    ]),
]
build_tc_sheet("☁ Services – NOAA", "1565C0", "1565C0", noaa_sections)


# ═══════════════════════════════════════════════════════════════════
#  SHEET 5 – CME PROCESSOR
# ═══════════════════════════════════════════════════════════════════
cme_serv_sections = [
    ("CMEProcessor – get_full_cme_package", [
        ("CM-001","CME Svc","Returns latest 10 CME events","NASA API key valid","processor.get_full_cme_package()","status='success'; total ≤ 10; each event has activityID, startTime, speed, latitude, longitude","","Pending","High",""),
        ("CM-002","CME Svc","Retries up to 3 times on 5xx errors","Simulate 502 responses","Mock response to return 502 twice then 200","Eventually succeeds, returns valid data","","Pending","Medium",""),
        ("CM-003","CME Svc","Raises exception when all retries fail","Simulate persistent failure","Mock all responses to return 503","Raises Exception('CME fetch failed:...')","","Pending","Medium",""),
        ("CM-004","CME Svc","Missing cmeAnalyses handled gracefully","CME item with no cmeAnalyses key","Pass event without cmeAnalyses","speed/latitude/longitude/type are all None; no crash","","Pending","High",""),
    ]),
    ("CMEProcessor – calculate_impact_probability", [
        ("CM-005","CME Svc","Halo CME + high speed + near-Earth longitude = High","No precond","calculate_impact_probability(1600, 10, 'halo')","Returns 'High'","","Pending","High",""),
        ("CM-006","CME Svc","Slow CME + large longitude offset = Low","No precond","calculate_impact_probability(300, 60, 'S')","Returns 'Low'","","Pending","High",""),
        ("CM-007","CME Svc","Medium CME returns Moderate","No precond","calculate_impact_probability(900, 20, 'C')","Returns 'Moderate'","","Pending","Medium",""),
        ("CM-008","CME Svc","None speed/longitude handled without crash","No precond","calculate_impact_probability(None, None, None)","Returns 'Low' without exception","","Pending","High",""),
    ]),
    ("CMEProcessor – get_latest_lasco_image", [
        ("CM-009","CME Svc","Downloads and saves LASCO GIF","SOHO reachable","processor.get_latest_lasco_image()","Returns path string; file exists at that path","","Pending","High",""),
        ("CM-010","CME Svc","Raises exception when SOHO unreachable","Simulate SOHO 404","Mock GET to return 404","Raises Exception('CME Image Error:...')","","Pending","Medium",""),
    ]),
]
build_tc_sheet("🛸 Services – CME", "E65100", "E65100", cme_serv_sections)


# ═══════════════════════════════════════════════════════════════════
#  SHEET 6 – SOLAR WIND SERVICE
# ═══════════════════════════════════════════════════════════════════
wind_sections = [
    ("SolarWindFetcher – _process_solar_wind", [
        ("SW-001","Wind Svc","Valid response parsed to list of dicts","Mock valid NOAA plasma response","Call _process_solar_wind(mock_response)","Returns list with time_tag, speed, density, temperature; all float","","Pending","High",""),
        ("SW-002","Wind Svc","Exception response returns empty list","Pass Exception object","_process_solar_wind(Exception('fail'))","Returns []","","Pending","High",""),
        ("SW-003","Wind Svc","Malformed rows skipped, valid rows kept","Mix valid+invalid rows in data","Pass mock response with one row having non-numeric speed","Invalid row skipped; valid rows present","","Pending","Medium",""),
        ("SW-004","Wind Svc","Returns ≤ 200 data points","Mock response with 500 rows","_process_solar_wind(large_response)","len(result) ≤ 200","","Pending","Medium",""),
    ]),
    ("SolarWindFetcher – _process_imf", [
        ("SW-005","Wind Svc","Valid response parsed to bx, by, bz, bt","Mock valid NOAA mag response","Call _process_imf(mock_response)","Returns list with bx, by, bz, bt as floats","","Pending","High",""),
        ("SW-006","Wind Svc","Exception response returns empty list","Pass Exception","_process_imf(Exception('fail'))","Returns []","","Pending","High",""),
    ]),
    ("SolarWindFetcher – async methods", [
        ("SW-007","Wind Svc","get_solar_wind_data returns non-empty list","NOAA reachable","await SolarWindFetcher.get_solar_wind_data()","Non-empty list of dicts","","Pending","High",""),
        ("SW-008","Wind Svc","get_imf_data returns non-empty list","NOAA reachable","await SolarWindFetcher.get_imf_data()","Non-empty list of dicts","","Pending","High",""),
        ("SW-009","Wind Svc","get_all_solar_wind_data fetches both concurrently","NOAA reachable","await SolarWindFetcher.get_all_solar_wind_data()","Dict with solar_wind and imf keys both non-empty","","Pending","High",""),
        ("SW-010","Wind Svc","get_all_solar_wind_data handles partial failure","Simulate plasma endpoint down","Mock plasma to raise, mag normal","solar_wind == [], imf is non-empty","","Pending","Medium",""),
    ]),
]
build_tc_sheet("💨 Services – Solar Wind", "004D40", "004D40", wind_sections)


# ═══════════════════════════════════════════════════════════════════
#  SHEET 7 – SEP SERVICE
# ═══════════════════════════════════════════════════════════════════
sep_sections = [
    ("SEPFetcher – _process_proton_flux", [
        ("SE-001","SEP Svc","Valid response parsed to list with time_tag, flux, energy, satellite","Mock NOAA proton JSON","_process_proton_flux(mock_response)","Non-empty list; all items have required keys","","Pending","High",""),
        ("SE-002","SEP Svc","Exception returns empty list","Pass Exception","_process_proton_flux(Exception())","Returns []","","Pending","High",""),
        ("SE-003","SEP Svc","Returns ≤ 300 items","Mock large response","_process_proton_flux(300+ row response)","len(result) ≤ 300","","Pending","Low",""),
    ]),
    ("SEPFetcher – _process_alerts", [
        ("SE-004","SEP Svc","SEP-related alerts filtered from full alert list","Mock alert list with mixed types","_process_alerts(mixed alerts response)","Only SEP/PROTON/ELECTRON/RADIATION keyword alerts returned","","Pending","High",""),
        ("SE-005","SEP Svc","risk_level='severe' when WA product_id present","Mock alert with WA product_id","_process_alerts(wa_alert_response)","risk_level == 'severe'","","Pending","High",""),
        ("SE-006","SEP Svc","risk_level='quiet' when no alerts","Mock empty alert list","_process_alerts(empty_response)","risk_level == 'quiet', alerts == []","","Pending","High",""),
        ("SE-007","SEP Svc","Alert messages truncated to 300 chars","Mock alert with 1000-char message","_process_alerts(long_message_response)","len(alert.message) ≤ 300","","Pending","Medium",""),
    ]),
    ("SEPFetcher – _calculate_radiation_risk", [
        ("SE-008","SEP Svc","avg_flux > 1000 returns severe crew risk","High flux data","_calculate_radiation_risk(very_high_flux_list)","crew='severe', satellite='high', deep_space='extreme'","","Pending","High",""),
        ("SE-009","SEP Svc","avg_flux 10–100 returns moderate crew risk","Medium flux data","_calculate_radiation_risk(medium_flux)","crew='moderate', satellite='low', deep_space='high'","","Pending","High",""),
        ("SE-010","SEP Svc","Empty proton list returns lowest risk","Empty list","_calculate_radiation_risk([])","crew='low', satellite='low', deep_space='moderate'","","Pending","Medium",""),
    ]),
    ("SEPFetcher – async methods", [
        ("SE-011","SEP Svc","get_particle_flux_data returns proton & electron","NOAA reachable","await SEPFetcher.get_particle_flux_data()","Dict with proton and electron lists","","Pending","High",""),
        ("SE-012","SEP Svc","get_all_sep_data returns all 3 keys","NOAA reachable","await SEPFetcher.get_all_sep_data()","particle_flux, alerts, radiation_risk all present","","Pending","High",""),
    ]),
]
build_tc_sheet("☢ Services – SEP", "B71C1C", "B71C1C", sep_sections)


# ═══════════════════════════════════════════════════════════════════
#  SHEET 8 – AI INFERENCE API
# ═══════════════════════════════════════════════════════════════════
ai_sections = [
    ("POST /ai/chat", [
        ("AI-001","AI Chat","Text-only chat returns valid response","Backend running","POST /ai/chat; form: message='hello'","200 OK; JSON with response string, surya_data, source, annotated_image=null, regions=[]","","Pending","High",""),
        ("AI-002","AI Chat","Chat with image upload processes image","Backend running; valid image file","POST /ai/chat with message + image file","200 OK; annotated_image may be non-null; regions may be non-empty","","Pending","High",""),
        ("AI-003","AI Chat","Invalid image returns 422","Backend running","POST /ai/chat with corrupt file as image","422 Unprocessable Entity; detail: 'Could not read the uploaded image.'","","Pending","High",""),
        ("AI-004","AI Chat","Missing message field returns 422","Backend running","POST /ai/chat without message field","422 Unprocessable Entity","","Pending","Medium",""),
    ]),
    ("POST /ai/flare-predict", [
        ("AI-005","AI Predict","Live data prediction returns complete response","Live GOES/NOAA reachable","POST /ai/flare-predict {use_live_data:true}","All FlareResponse fields non-empty; predicted_class is A/B/C/M/X","","Pending","High",""),
        ("AI-006","AI Predict","Manual flux override works","No precond","POST /ai/flare-predict {use_live_data:false, flux_window:[...12 values...]}","Valid prediction using provided flux","","Pending","High",""),
        ("AI-007","AI Predict","include_explanation=false returns empty reasoning","Backend running","POST /ai/flare-predict {include_explanation:false}","reasoning is empty string","","Pending","Medium",""),
        ("AI-008","AI Predict","timestamp is valid ISO format","Backend running","POST /ai/flare-predict","timestamp parses as ISO datetime with Z suffix","","Pending","Medium",""),
        ("AI-009","AI Predict","confidence field is non-empty string","Backend running","POST /ai/flare-predict","confidence field is non-empty","","Pending","Medium",""),
        ("AI-010","AI Predict","onset_window_minutes is positive integer","Backend running","POST /ai/flare-predict","onset_window_minutes > 0","","Pending","Medium",""),
    ]),
]
build_tc_sheet("🤖 AI – Inference", "4A148C", "4A148C", ai_sections)


# ═══════════════════════════════════════════════════════════════════
#  SHEET 9 – FRONTEND PAGES
# ═══════════════════════════════════════════════════════════════════
fe_pages_sections = [
    ("Home Page (/)", [
        ("FP-001","Home","Home page renders STELAR hero text","Frontend running","Navigate to http://localhost:3000","Page shows 'STELAR' heading and subtitle text","","Pending","High",""),
        ("FP-002","Home","Starfield background canvas renders","Frontend running","Navigate to home","Canvas/Starfield element visible in DOM","","Pending","Medium",""),
        ("FP-003","Home","Navbar links are visible","Frontend running","Navigate to home","All nav items rendered; no console errors","","Pending","High",""),
    ]),
    ("Solar Flare Page (/solar-flare)", [
        ("FP-004","Solar Flare Page","Page loads with HMI Magnetogram card","Full stack running","Navigate to /solar-flare","Magnetogram image renders; C/M/X probability badges visible","","Pending","High",""),
        ("FP-005","Solar Flare Page","GOES X-ray Flux chart renders","Full stack running","Navigate to /solar-flare; scroll to GOES card","Plotly chart with GOES-16 and GOES-17 traces visible","","Pending","High",""),
        ("FP-006","Solar Flare Page","AIA EUV wavelength selector works (94Å,131Å,171Å,193Å)","Full stack running","Click each wavelength button; observe image","Image updates for each wavelength selection","","Pending","High",""),
        ("FP-007","Solar Flare Page","Recent Flare Events table shows data","Full stack running","Scroll to Recent Flare Events card","Table with Class, Start, Peak, End, Region columns populated","","Pending","High",""),
        ("FP-008","Solar Flare Page","Flare class X badge shows red color","Active X-class event in data","View flare event log","X-class entries show red badge","","Pending","Medium",""),
    ]),
    ("CME Page (/cme)", [
        ("FP-009","CME Page","CME Velocity card shows speed in km/s","Full stack running","Navigate to /cme; scroll to CME Velocity card","Speed number in orange text + 'km/s' label","","Pending","High",""),
        ("FP-010","CME Page","Magnetic Structure card shows type, lat, lon, half angle","Full stack running","Scroll to Magnetic Structure card","All 4 fields displayed; N/A shown when null","","Pending","High",""),
        ("FP-011","CME Page","Impact Probability bar chart shows last 10 events","Full stack running","Scroll to Impact Probability card","Bar chart with High/Moderate/Low breakdown visible","","Pending","High",""),
        ("FP-012","CME Page","CME Coronagraph Image shows LASCO GIF","Full stack + SOHO reachable","Scroll to coronagraph card","GIF from LASCO loads; no broken image icon","","Pending","High",""),
        ("FP-013","CME Page","CME Event Log table shows 10 rows","Full stack running","Scroll to CME Event Log card","Table with ID, Time, Speed, Type, Location, Risk columns","","Pending","High",""),
    ]),
    ("SEP Page (/sep)", [
        ("FP-014","SEP Page","Proton Flux stat card displays current flux","Full stack running","Navigate to /sep","Proton Flux (≥10 MeV) card shows numeric value + pfu unit","","Pending","High",""),
        ("FP-015","SEP Page","Radiation Risk by Mission Type section shows 3 cards","Full stack running","Navigate to /sep; wait for data","Crew, Satellite, Deep Space risk cards render with color-coded levels","","Pending","High",""),
        ("FP-016","SEP Page","SEP event warning shown when flux > 10 pfu","Simulated high flux data","Navigate to /sep with high proton flux","Red '⚠️ SEP Event in progress' text shown","","Pending","Medium",""),
        ("FP-017","SEP Page","Data refreshes every 5 minutes","Full stack running","Wait 5 minutes on /sep page","Last updated timestamp changes after 5 min","","Pending","Low",""),
    ]),
    ("Solar Wind Page (/solar-wind)", [
        ("FP-018","Wind Page","Solar Wind Speed stat card shows km/s","Full stack running","Navigate to /solar-wind","Speed in cyan text with km/s unit; avg + status shown","","Pending","High",""),
        ("FP-019","Wind Page","IMF Bz < -10 nT shows red storm risk warning","Simulate Bz = -15","Navigate to /solar-wind with negative Bz","Bz card shows red '⚠️ Strong southward — storm risk' text","","Pending","High",""),
        ("FP-020","Wind Page","Error banner shows Retry button on fetch failure","Simulate API 500","Navigate to /solar-wind with backend returning 500","Red error banner and Retry button visible","","Pending","High",""),
    ]),
]
build_tc_sheet("🌐 Frontend – Pages", "1B5E20", "1B5E20", fe_pages_sections)


# ═══════════════════════════════════════════════════════════════════
#  SHEET 10 – FRONTEND COMPONENTS
# ═══════════════════════════════════════════════════════════════════
fe_comp_sections = [
    ("GOESFluxChart Component", [
        ("FC-001","GOESFluxChart","Renders Plotly chart with two traces","GOES data available","Mount GOESFluxChart","Chart has 'GOES-16 (Primary)' and 'GOES-17 (Secondary)' legend items","","Pending","High",""),
        ("FC-002","GOESFluxChart","Y-axis is logarithmic","GOES data available","Inspect chart y-axis","type='log' applied","","Pending","Medium",""),
        ("FC-003","GOESFluxChart","Loading state shows loading text","Simulate delayed fetch","Mount component; observe before data arrives","'Loading GOES X-Ray Data...' text shown","","Pending","Medium",""),
        ("FC-004","GOESFluxChart","No data state shows 'No flux data available'","Simulate empty response","Mock API returns {primary:[],secondary:[]}","'No flux data available.' text shown","","Pending","Medium",""),
    ]),
    ("FlareEventLog Component", [
        ("FC-005","FlareLog","Renders table with Class, Start, Peak, End, Region columns","Flare data available","Mount FlareEventLog","All 5 column headers visible","","Pending","High",""),
        ("FC-006","FlareLog","X-class shows red badge","X-class flare in data","Mount with X-class data","bg-red-500/20 + text-red-300 applied","","Pending","High",""),
        ("FC-007","FlareLog","M-class shows orange badge","M-class flare in data","Mount with M-class data","bg-orange-500/20 badge applied","","Pending","Medium",""),
        ("FC-008","FlareLog","Shows last 16 events in reverse order","More than 16 flares in API response","Mount FlareEventLog","Exactly 16 rows; most recent first","","Pending","Medium",""),
        ("FC-009","FlareLog","Empty data shows 'No flare events found'","Empty flare array","Mock API returns []","'No flare events found.' message shown","","Pending","Medium",""),
    ]),
    ("CMECards Components", [
        ("FC-010","CMECards","CMEVelocityContent shows speed number","CME data available","Mount CMEVelocityContent","Speed in orange large text; 'km/s' label","","Pending","High",""),
        ("FC-011","CMECards","CMEImpactContent shows High/Moderate/Low bar chart","CME data available","Mount CMEImpactContent","Three horizontal progress bars with counts","","Pending","High",""),
        ("FC-012","CMECards","CMEImageContent renders <img> with CME image URL","Backend running","Mount CMEImageContent","img src points to /space-weather/cme/image","","Pending","High",""),
        ("FC-013","CMECards","CMEEventLog shows 6-column table","CME data available","Mount CMEEventLog","ID, Time, Speed, Type, Location, Risk columns","","Pending","High",""),
    ]),
    ("GlassCard & ServiceCards", [
        ("FC-014","GlassCard","Renders children inside glass-styled div","React env","Mount GlassCard with child text","glass class applied; child text visible","","Pending","Low",""),
    ]),
]
build_tc_sheet("🧩 Frontend – Components", "37474F", "37474F", fe_comp_sections)


# ═══════════════════════════════════════════════════════════════════
#  SHEET 11 – FRONTEND LIB/API.TS
# ═══════════════════════════════════════════════════════════════════
api_lib_sections = [
    ("lib/api.ts – fetch functions", [
        ("LI-001","API Lib","getMagnetogramImageUrl returns correct URL string","NEXT_PUBLIC_API_URL set","Call getMagnetogramImageUrl()","Returns 'http://localhost:8000/space-weather/magnetogram/image'","","Pending","High",""),
        ("LI-002","API Lib","getGoesXrayFlux fetches /noaa/goes-xray","Backend running","await getGoesXrayFlux()","Returns {primary, secondary}","","Pending","High",""),
        ("LI-003","API Lib","getGoesXrayFlux throws on non-ok response","Simulate 500","Mock fetch to return status 500","Throws Error('GOES fetch failed')","","Pending","High",""),
        ("LI-004","API Lib","getSolarFlares fetches /space-weather/flares","Backend running","await getSolarFlares()","Returns array of flare objects","","Pending","High",""),
        ("LI-005","API Lib","getAIAImageUrl maps wavelength label to code correctly","No precond","getAIAImageUrl('171Å')","Returns URL with wavelength=0171","","Pending","High",""),
        ("LI-006","API Lib","getCMEData fetches /space-weather/cme/full","Backend running","await getCMEData()","Returns {status, total, cme_events}","","Pending","High",""),
        ("LI-007","API Lib","getCMEImageUrl returns correct URL","NEXT_PUBLIC_API_URL set","getCMEImageUrl()","Returns '.../ space-weather/cme/image'","","Pending","Medium",""),
        ("LI-008","API Lib","getSolarWindData fetches /wind/speed","Backend running","await getSolarWindData()","Returns {status, data}","","Pending","High",""),
        ("LI-009","API Lib","getIMFData fetches /wind/imf","Backend running","await getIMFData()","Returns {status, data}","","Pending","High",""),
        ("LI-010","API Lib","getAllSEPData fetches /sep/all","Backend running","await getAllSEPData()","Returns {status, particle_flux, alerts, radiation_risk}","","Pending","High",""),
    ]),
]
build_tc_sheet("🔗 Frontend – API lib", "1A237E", "1A237E", api_lib_sections)


# ═══════════════════════════════════════════════════════════════════
#  SHEET 12 – NON-FUNCTIONAL
# ═══════════════════════════════════════════════════════════════════
nf_sections = [
    ("Performance", [
        ("NF-001","Perf","Cached magnetogram endpoint responds in < 2 sec","Cache exists","GET /magnetogram/image; measure response time","Response time < 2000 ms","","Pending","High",""),
        ("NF-002","Perf","GOES X-ray endpoint responds in < 5 sec","NOAA reachable","GET /noaa/goes-xray; measure response time","Response time < 5000 ms","","Pending","High",""),
        ("NF-003","Perf","Frontend page loads in < 3 sec","Full stack running","Open /solar-flare; measure LCP","LCP ≤ 3 sec","","Pending","Medium",""),
        ("NF-004","Perf","Concurrent requests (10) to /wind/all handled without error","Backend running","Fire 10 simultaneous requests","All 10 return 200 OK","","Pending","Medium",""),
    ]),
    ("Security", [
        ("NF-005","Security","CORS allows frontend origin correctly","Frontend on :3000, backend on :8000","Open browser DevTools; check CORS headers","Access-Control-Allow-Origin: * present","","Pending","High",""),
        ("NF-006","Security","NASA API key not exposed in frontend bundle","Frontend built","Search compiled JS bundle for API key string","No API key found in frontend assets","","Pending","High",""),
        ("NF-007","Security","SQL/command injection in message field not executed","Backend running","POST /ai/chat with message='; DROP TABLE'","Response processes as text; no server crash","","Pending","High",""),
    ]),
    ("Error Handling", [
        ("NF-008","Error","Backend returns JSON error body not HTML on 500","Backend running","Trigger internal error","Content-Type is application/json; detail field present","","Pending","High",""),
        ("NF-009","Error","Frontend shows error banner when backend is unreachable","Stop backend","Navigate to /solar-wind","Red error banner with Retry button shown","","Pending","High",""),
        ("NF-010","Error","Retry button re-fetches data","Backend was down then restarted","Click Retry after backend returns","Data loads successfully","","Pending","High",""),
        ("NF-011","Error","Loading skeleton shown while data is fetching","Slow network","Navigate to /sep; throttle network to Slow 3G","Skeleton animation visible before data arrives","","Pending","Medium",""),
    ]),
    ("Accessibility & UX", [
        ("NF-012","A11y","Images have alt attributes","Frontend running","Inspect img tags on all pages","All <img> have descriptive alt text","","Pending","Medium",""),
        ("NF-013","A11y","Tab navigation works across sidebar links","Frontend running","Press Tab through sidebar","Focus visible on each sidebar icon in order","","Pending","Medium",""),
        ("NF-014","A11y","Color contrast ratio ≥ 4.5:1 for body text","Frontend running","Run Lighthouse accessibility audit","Contrast ratio ≥ 4.5:1 on primary text","","Pending","Medium",""),
    ]),
]
build_tc_sheet("🔒 Non-Functional", "546E7A", "546E7A", nf_sections)


# ═══════════════════════════════════════════════════════════════════
#  SHEET 13 – INTEGRATION / END-TO-END
# ═══════════════════════════════════════════════════════════════════
e2e_sections = [
    ("End-to-End Flows", [
        ("E2E-001","E2E","Solar Flare full flow: open page → see live magnetogram image","Full stack running","(1) Start backend (2) Start frontend (3) Navigate to /solar-flare","Magnetogram image renders with probability badges; no console errors","","Pending","High",""),
        ("E2E-002","E2E","GOES chart updates when backend data changes","Full stack running","(1) Navigate to /solar-flare (2) Note current chart (3) Wait for refresh","Chart data reflects latest GOES readings","","Pending","High",""),
        ("E2E-003","E2E","CME full flow: all 5 cards load data","Full stack running","Navigate to /cme; scroll through all 5 cards","Each card shows data or appropriate empty state","","Pending","High",""),
        ("E2E-004","E2E","SEP page full flow: proton flux + alerts + risk levels","Full stack running","Navigate to /sep","Proton flux, Radiation Risk Level, and Mission Type cards all populated","","Pending","High",""),
        ("E2E-005","E2E","Solar Wind full flow: speed + IMF cards displayed","Full stack running","Navigate to /solar-wind","Speed, Plasma Density, Bz shown with correct units","","Pending","High",""),
        ("E2E-006","E2E","AIA wavelength switch updates image in < 3 sec","Full stack running","(1) Navigate to /solar-flare (2) Click 94Å → 193Å → 131Å → 171Å","Image updates for each wavelength within 3 sec","","Pending","High",""),
        ("E2E-007","E2E","Backend restart recovers without frontend restart","Full stack running","(1) Stop backend (2) Note error banner on FE (3) Restart backend (4) Click Retry","Data loads after retry; page returns to normal state","","Pending","High",""),
        ("E2E-008","E2E","NASA API rate limit graceful handling (50 req/day)","Near rate limit","Make repeated calls to /flares and /cme/full","After limit hit, error logged; frontend shows error state not crash","","Pending","Medium","Requires monitoring rate limit count"),
        ("E2E-009","E2E","CORS allows frontend fetch from Codespaces forwarded URL","Codespace env","(1) Deploy to Codespace (2) Set NEXT_PUBLIC_API_URL to forwarded URL (3) Open FE","All API calls succeed; no CORS errors in console","","Pending","High",""),
        ("E2E-010","E2E","AI flare-predict endpoint integrates with live GOES+NOAA data","Full stack + live APIs","POST /ai/flare-predict {use_live_data:true}","predicted_class based on real current data; goes_peak_flux matches recent GOES reading","","Pending","Medium",""),
    ]),
]
build_tc_sheet("🔄 Integration E2E", "004D40", "004D40", e2e_sections)


# ── freeze panes on all TC sheets ───────────────────────────────────
for ws in wb.worksheets:
    if ws.title != "📋 Index":
        ws.freeze_panes = "A3"

# ── save ────────────────────────────────────────────────────────────
out = r"c:\Users\sruja\Desktop\Analytics-Dashboard-for-SSA\SSA_Dashboard_Test_Cases.xlsx"
wb.save(out)
print(f"✅  Excel saved to: {out}")
